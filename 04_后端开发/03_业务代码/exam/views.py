from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from django.db import transaction
from .models import Question, Paper, Exam, ExamAnswer, ScoreRecord
from .serializers import QuestionSerializer, PaperSerializer, ExamSerializer, ScoreRecordSerializer
from edu.models import Student, Teacher, Subject



class QuestionViewSet(viewsets.ModelViewSet):
    """题目管理视图"""
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['subject', 'type', 'difficulty', 'status']
    search_fields = ['content']

    def get_queryset(self):
        return Question.objects.select_related('subject', 'created_by').all().order_by('-id')

    def _build_question_workbook(self, questions=None, template=False):
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise RuntimeError('当前环境缺少 openpyxl，无法导出 Excel') from exc

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '题库'
        headers = ['科目编码', '章节', '题型', '题目内容', '选项', '答案', '解析', '难度', '分值', '状态']
        worksheet.append(headers)

        if template:
            worksheet.append([
                'MATH',
                '函数基础',
                'single',
                '已知 f(x)=x+1，f(2)=?',
                'A.1\nB.2\nC.3\nD.4',
                'C',
                '把 x=2 代入即可',
                'easy',
                5,
                'true',
            ])
            worksheet.append([
                'MATH',
                '函数基础',
                'essay',
                '请简述一次函数图像的特点',
                '',
                '略',
                '围绕斜率和截距作答',
                'medium',
                10,
                'true',
            ])

        for item in questions or []:
            options = ''
            if isinstance(item.options, dict):
                options = '\n'.join([f'{key}. {value}' for key, value in item.options.items()])
            worksheet.append([
                item.subject.code,
                item.chapter,
                item.type,
                item.content,
                options,
                item.answer,
                item.analysis,
                item.difficulty,
                item.score,
                'true' if item.status else 'false',
            ])

        return workbook

    def _parse_options(self, raw_text):
        lines = [line.strip() for line in str(raw_text or '').splitlines() if line.strip()]
        options = {}
        for index, line in enumerate(lines):
            key = chr(65 + index)
            options[key] = line.replace(f'{key}.', '').replace(f'{key}、', '').strip()
        return options

    def _load_import_rows(self, upload):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError('当前环境缺少 openpyxl，无法导入 Excel')

        workbook = load_workbook(upload)
        worksheet = workbook.active
        return list(worksheet.iter_rows(min_row=2, values_only=True))

    def _validate_import_rows(self, rows):
        preview = []
        created_count = 0
        updated_count = 0
        errors = []

        for row_index, row in enumerate(rows, start=2):
            subject_code, chapter, qtype, content, options_text, answer, analysis, difficulty, score, enabled = (list(row[:10]) + [None] * 10)[:10]
            if not content:
                continue

            row_errors = []
            subject = Subject.objects.filter(code=str(subject_code or '').strip()).first()
            if not subject:
                row_errors.append(f'科目编码不存在: {subject_code}')

            payload = {
                'subject': subject,
                'chapter': str(chapter or '').strip(),
                'type': str(qtype or 'single').strip(),
                'content': str(content).strip(),
                'options': self._parse_options(options_text),
                'answer': str(answer or '').strip(),
                'analysis': str(analysis or '').strip(),
                'difficulty': str(difficulty or 'medium').strip(),
                'score': int(score or 5),
                'status': str(enabled).strip().lower() not in ['false', '0', 'no'],
            }

            if payload['type'] not in ['single', 'multiple', 'blank', 'essay']:
                row_errors.append(f'题型不支持: {payload["type"]}')
            if payload['difficulty'] not in ['easy', 'medium', 'hard']:
                row_errors.append(f'难度不支持: {payload["difficulty"]}')
            if not payload['answer']:
                row_errors.append('答案不能为空')
            if payload['score'] <= 0:
                row_errors.append('分值必须大于 0')

            exists = bool(subject and Question.objects.filter(subject=subject, content=payload['content']).exists())
            action = '更新' if exists else '新增'
            if exists:
                updated_count += 1
            else:
                created_count += 1

            preview.append({
                'row': row_index,
                'subject_code': subject_code,
                'chapter': payload['chapter'],
                'content': payload['content'],
                'type': payload['type'],
                'options_text': str(options_text or ''),
                'answer': payload['answer'],
                'analysis': payload['analysis'],
                'difficulty': payload['difficulty'],
                'score': payload['score'],
                'status': payload['status'],
                'action': action,
                'errors': row_errors,
            })

            if row_errors:
                errors.extend([f'第 {row_index} 行{error}' for error in row_errors])

        return {
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'preview': preview,
        }

    @action(detail=False, methods=['get'])
    def export(self, request):
        questions = self.filter_queryset(self.get_queryset())
        workbook = self._build_question_workbook(questions=questions)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=questions.xlsx'
        workbook.save(response)
        return response

    @action(detail=False, methods=['get'])
    def template(self, request):
        workbook = self._build_question_workbook(template=True)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=question_import_template.xlsx'
        workbook.save(response)
        return response

    @action(detail=False, methods=['post'])
    def import_preview(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': '请上传 Excel 文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = self._load_import_rows(upload)
        except RuntimeError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if not rows:
            return Response({'error': '导入文件为空'}, status=status.HTTP_400_BAD_REQUEST)

        result = self._validate_import_rows(rows)
        return Response(result)

    @transaction.atomic
    @action(detail=False, methods=['post'])
    def import_file(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': '请上传 Excel 文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = self._load_import_rows(upload)
        except RuntimeError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if not rows:
            return Response({'error': '导入文件为空'}, status=status.HTTP_400_BAD_REQUEST)

        preview_result = self._validate_import_rows(rows)
        blocking_errors = [item for item in preview_result['preview'] if item['errors']]
        if blocking_errors:
            return Response({
                'error': '导入文件存在校验错误，请先修正后再导入',
                **preview_result,
            }, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        update_count = 0

        for row_index, row in enumerate(rows, start=2):
            subject_code, chapter, qtype, content, options_text, answer, analysis, difficulty, score, enabled = row[:10]
            if not content:
                continue

            try:
                subject = Subject.objects.get(code=str(subject_code).strip())
            except Subject.DoesNotExist:
                continue

            payload = {
                'subject': subject,
                'chapter': str(chapter or '').strip(),
                'type': str(qtype or 'single').strip(),
                'content': str(content).strip(),
                'options': self._parse_options(options_text),
                'answer': str(answer or '').strip(),
                'analysis': str(analysis or '').strip(),
                'difficulty': str(difficulty or 'medium').strip(),
                'score': int(score or 5),
                'status': str(enabled).strip().lower() not in ['false', '0', 'no'],
            }

            if not payload['answer']:
                continue

            question, created = Question.objects.update_or_create(
                subject=subject,
                content=payload['content'],
                defaults={
                    **payload,
                    'created_by': request.user,
                }
            )
            if created:
                created_count += 1
            else:
                update_count += 1

        return Response({
            'message': '导入完成',
            'created': created_count,
            'updated': update_count,
            'errors': [],
        })


class PaperViewSet(viewsets.ModelViewSet):
    """试卷管理视图"""
    queryset = Paper.objects.all()
    serializer_class = PaperSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['subject', 'status']


class ExamViewSet(viewsets.ModelViewSet):
    """考试管理视图"""
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['edu_class', 'status']

    def get_permissions(self):
        if self.action == 'submit':
            return [permissions.AllowAny()]
        return super().get_permissions()

    def _get_user_phone(self, user):
        return getattr(user, 'phone', None) or getattr(user, 'username', None)

    def _get_role_codes(self, user):
        return set(user.roles.values_list('code', flat=True))

    def _get_role_names(self, user):
        return set(user.roles.values_list('name', flat=True))

    def _is_admin(self, user):
        role_codes = self._get_role_codes(user)
        role_names = self._get_role_names(user)
        return user.is_superuser or 'admin' in role_codes or '管理员' in role_names

    def get_queryset(self):
        queryset = Exam.objects.select_related('paper', 'edu_class').all().order_by('-start_time')
        user = getattr(self.request, 'user', None)
        if not user or not user.is_authenticated or self._is_admin(user):
            return queryset

        user_phone = self._get_user_phone(user)
        role_codes = self._get_role_codes(user)
        role_names = self._get_role_names(user)
        filters = Q(pk__in=[])

        if 'teacher' in role_codes or '教师' in role_names:
            teacher_ids = Teacher.objects.filter(phone=user_phone).values_list('id', flat=True)
            filters |= Q(edu_class__teacher_id__in=teacher_ids)

        if 'student' in role_codes or '学生' in role_names:
            student_ids = Student.objects.filter(phone=user_phone).values_list('id', flat=True)
            filters |= Q(
                edu_class__class_students__student_id__in=student_ids,
                edu_class__class_students__status='studying',
            )

        if 'parent' in role_codes or '家长' in role_names:
            filters |= Q(
                edu_class__class_students__student__parent_phone=user_phone,
                edu_class__class_students__status='studying',
            )

        return queryset.filter(filters).distinct()

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        """发布考试"""
        exam = self.get_object()
        exam.status = 'ongoing' if exam.start_time <= timezone.now() <= exam.end_time else 'pending'
        exam.save()
        return Response({'message': '考试已发布'})

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """提交考试答案并生成成绩记录"""
        exam = self.get_object()
        answers = request.data.get('answers') or {}
        student_id = request.data.get('student_id')

        if not student_id:
            return Response({'error': '缺少 student_id 参数'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({'error': '学生不存在'}, status=status.HTTP_404_NOT_FOUND)

        if exam.edu_class and not exam.edu_class.class_students.filter(student=student, status='studying').exists():
            return Response({'error': '当前学生不在本场考试所属班级中'}, status=status.HTTP_403_FORBIDDEN)

        questions = list(exam.paper.questions.all().order_by('paperquestion__sort', 'id'))
        if not questions:
            return Response({'error': '当前考试未配置题目'}, status=status.HTTP_400_BAD_REQUEST)

        total_score = 0
        earned_score = 0

        for index, question in enumerate(questions):
            raw_answer = answers.get(str(index), answers.get(index, ''))
            submitted_answer = '' if raw_answer is None else str(raw_answer)
            total_score += question.score

            normalized_expected = str(question.answer).strip()
            normalized_submitted = submitted_answer.strip()
            is_auto = question.type in ['single', 'multiple', 'blank']
            score = question.score if is_auto and normalized_submitted == normalized_expected else 0
            earned_score += score

            ExamAnswer.objects.update_or_create(
                exam=exam,
                student_id=student_id,
                question=question,
                defaults={
                    'answer': submitted_answer,
                    'score': score,
                    'is_auto': is_auto,
                },
            )

        score_record, _ = ScoreRecord.objects.update_or_create(
            exam=exam,
            student_id=student_id,
            defaults={
                'total_score': total_score,
                'score': earned_score,
            },
        )

        ranked_scores = list(ScoreRecord.objects.filter(exam=exam).order_by('-score', 'created_at'))
        for rank, item in enumerate(ranked_scores, start=1):
            if item.rank != rank:
                item.rank = rank
                item.save(update_fields=['rank'])
        score_record.refresh_from_db()

        return Response({
            'message': '提交成功',
            'student_id': student_id,
            'score': score_record.score,
            'total_score': score_record.total_score,
            'rank': score_record.rank,
        })

    @action(detail=True, methods=['get'])
    def essay_answers(self, request, pk=None):
        """获取需要批改的主观题列表"""
        exam = self.get_object()
        
        # 获取所有主观题答案
        essay_answers = ExamAnswer.objects.filter(
            exam=exam,
            question__type='essay',
            is_auto=False
        ).select_related('question', 'question__subject')
        
        # 构建响应数据
        data = []
        for answer in essay_answers:
            data.append({
                'id': answer.id,
                'student_id': answer.student_id,
                'question_id': answer.question.id,
                'question_content': answer.question.content,
                'question_score': answer.question.score,
                'answer': answer.answer,
                'score': answer.score,
                'created_at': answer.created_at
            })
        
        return Response({'answers': data})

    @action(detail=True, methods=['post'])
    def grade_essay(self, request, pk=None):
        """批改主观题"""
        exam = self.get_object()
        answer_id = request.data.get('answer_id')
        score = request.data.get('score')
        
        if not answer_id or score is None:
            return Response({'error': '缺少必要参数'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            answer = ExamAnswer.objects.get(id=answer_id, exam=exam)
        except ExamAnswer.DoesNotExist:
            return Response({'error': '答案不存在'}, status=status.HTTP_404_NOT_FOUND)
        
        # 更新分数
        answer.score = score
        answer.is_auto = False
        answer.save()
        
        # 更新总分
        student_id = answer.student_id
        exam_answers = ExamAnswer.objects.filter(exam=exam, student_id=student_id)
        total_score = sum(answer.score or 0 for answer in exam_answers)
        
        score_record, _ = ScoreRecord.objects.update_or_create(
            exam=exam,
            student_id=student_id,
            defaults={
                'score': total_score
            }
        )
        
        # 更新排名
        ranked_scores = list(ScoreRecord.objects.filter(exam=exam).order_by('-score', 'created_at'))
        for rank, item in enumerate(ranked_scores, start=1):
            if item.rank != rank:
                item.rank = rank
                item.save(update_fields=['rank'])
        score_record.refresh_from_db()
        
        return Response({
            'message': '批改成功',
            'answer_id': answer_id,
            'score': score,
            'student_id': student_id,
            'total_score': total_score,
            'rank': score_record.rank
        })


class ScoreRecordViewSet(viewsets.ReadOnlyModelViewSet):
    """成绩记录视图"""
    queryset = ScoreRecord.objects.all()
    serializer_class = ScoreRecordSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['exam', 'student_id']

    def _get_user_phone(self, user):
        return getattr(user, 'phone', None) or getattr(user, 'username', None)

    def _get_role_codes(self, user):
        return set(user.roles.values_list('code', flat=True))

    def _get_role_names(self, user):
        return set(user.roles.values_list('name', flat=True))

    def _is_admin(self, user):
        role_codes = self._get_role_codes(user)
        role_names = self._get_role_names(user)
        return user.is_superuser or 'admin' in role_codes or '管理员' in role_names

    def get_queryset(self):
        queryset = super().get_queryset().select_related('exam', 'exam__edu_class', 'exam__edu_class__teacher')
        user = self.request.user

        if self._is_admin(user):
            return queryset

        user_phone = self._get_user_phone(user)
        role_codes = self._get_role_codes(user)
        role_names = self._get_role_names(user)
        filters = Q(pk__in=[])

        if 'teacher' in role_codes or '教师' in role_names:
            teacher_ids = Teacher.objects.filter(phone=user_phone).values_list('id', flat=True)
            filters |= Q(exam__edu_class__teacher_id__in=teacher_ids)

        if 'student' in role_codes or '学生' in role_names:
            student_ids = list(Student.objects.filter(phone=user_phone).values_list('id', flat=True))
            filters |= Q(student_id__in=student_ids)

        if 'parent' in role_codes or '家长' in role_names:
            student_ids = list(Student.objects.filter(parent_phone=user_phone).values_list('id', flat=True))
            filters |= Q(student_id__in=student_ids)

        return queryset.filter(filters).distinct()

    @action(detail=False, methods=['get'])
    def by_student(self, request):
        """按学生查询"""
        queryset = self.filter_queryset(self.get_queryset())
        student_id = request.query_params.get('student_id')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
